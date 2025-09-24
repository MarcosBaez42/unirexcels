"""Utility to merge multiple Excel workbooks into a single file.

This module provides a command line interface and a callable helper
function that copies the first worksheet from every Excel file in a
folder into a consolidated workbook.  Each worksheet in the output is
named after the corresponding source file (without its extension).
"""
from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterator, List, Sequence

from copy import copy, deepcopy

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.cell import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

MAX_SHEET_NAME_LENGTH = 31
_INVALID_SHEET_CHARS = set("[]:*?/\\")
# Folder used when the script is executed without explicitly providing a
# ``source_directory``.  The value can be absolute or relative to the location
# of this script.  If the path does not exist at runtime the script falls back to
# using the directory that contains ``merge_excel_files.py``.
DEFAULT_SOURCE_DIRECTORY = Path("mis_excels/Reportes")

@dataclass(frozen=True)
class MergedSheet:
    """Represents the relationship between a source file and a sheet name."""

    source: Path
    sheet_name: str


def merge_excel_files(
    source_directory: Path,
    output_path: Path,
    pattern: str = "*.xlsx",
    recursive: bool = False,
    values_only: bool = False,
) -> List[MergedSheet]:
    """Merge Excel workbooks stored inside *source_directory* into one file.

    Parameters
    ----------
    source_directory:
        Folder that contains the files to merge.
    output_path:
        Path where the consolidated workbook should be written.  The parent
        directory will be created automatically if it does not exist.
    pattern:
        Glob expression used to filter the files inside ``source_directory``.
        Defaults to ``"*.xlsx"`` which will match the most common Excel
        workbooks.  The comparison is case-sensitive.
    recursive:
        When ``True`` the pattern is applied recursively.
    values_only:
        When ``True`` formulas in the source sheets are replaced by their
        last cached value.  When ``False`` formulas are preserved.

    Returns
    -------
    List[MergedSheet]
        A list describing which files were merged and the sheet names used
        in the output workbook.  The sheets are ordered according to the
        file names used for the merge.

    Raises
    ------
    FileNotFoundError
        If ``source_directory`` does not exist.
    ValueError
        If no Excel files matching ``pattern`` are found.
    """

    if not source_directory.exists():
        raise FileNotFoundError(f"Source directory '{source_directory}' does not exist")
    if not source_directory.is_dir():
        raise NotADirectoryError(f"'{source_directory}' is not a directory")

    resolved_output = output_path.resolve()

    candidates = _collect_excel_files(source_directory, pattern, recursive)
    files = [path for path in candidates if path.resolve() != resolved_output]

    if not files and not recursive:
        recursive_candidates = _collect_excel_files(source_directory, pattern, True)
        files = [path for path in recursive_candidates if path.resolve() != resolved_output]

    if not files:
        raise ValueError(
            f"No Excel files matching pattern '{pattern}' were found in '{source_directory}'"
        )

    workbook = Workbook()
    # Remove the default sheet created by openpyxl; it will be replaced with
    # the sheets copied from the source files.
    if workbook.sheetnames:
        workbook.remove(workbook.active)

    existing_names: set[str] = set()
    merged_sheets: List[MergedSheet] = []

    for file_path in files:
        sheet_title = _build_sheet_title(file_path.stem, existing_names)
        source_wb = load_workbook(filename=file_path, data_only=values_only)
        try:
            source_sheet = source_wb.worksheets[0]
            target_sheet = workbook.create_sheet(title=sheet_title)
            _copy_sheet_contents(source_sheet, target_sheet)
        finally:
            source_wb.close()
        merged_sheets.append(MergedSheet(source=file_path, sheet_name=sheet_title))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    return merged_sheets


def _collect_excel_files(directory: Path, pattern: str, recursive: bool) -> List[Path]:
    search_method = directory.rglob if recursive else directory.glob
    files = sorted(path for path in search_method(pattern) if path.is_file())
    return files


def _build_sheet_title(source_stem: str, existing_names: set[str]) -> str:
    base_name = _sanitize_sheet_base(source_stem)
    for candidate in _generate_sheet_name_candidates(base_name):
        if candidate not in existing_names:
            existing_names.add(candidate)
            return candidate
    raise RuntimeError("Unable to determine a unique sheet name")


def _sanitize_sheet_base(raw_name: str) -> str:
    sanitized = "".join(
        "_" if ch in _INVALID_SHEET_CHARS or ord(ch) < 32 else ch for ch in raw_name
    ).strip()
    sanitized = sanitized.rstrip("'")
    return sanitized or "Sheet"


def _generate_sheet_name_candidates(base_name: str) -> Iterator[str]:
    trimmed = base_name[:MAX_SHEET_NAME_LENGTH]
    if trimmed:
        yield trimmed
    else:
        yield "Sheet"

    counter = 2
    while True:
        suffix = f"_{counter}"
        trimmed_length = MAX_SHEET_NAME_LENGTH - len(suffix)
        prefix = base_name[:trimmed_length].rstrip()
        if not prefix:
            prefix = base_name[:trimmed_length]
        if not prefix:
            prefix = "Sheet"[:trimmed_length]
        candidate = f"{prefix}{suffix}"[:MAX_SHEET_NAME_LENGTH]
        if candidate:
            yield candidate
        counter += 1


def _copy_sheet_contents(source: Worksheet, target: Worksheet) -> None:
    if source.max_row == 1 and source.max_column == 1 and source.cell(1, 1).value is None:
        # Leave the sheet empty if the source is empty.
        return

    for row in source.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue

            column_index = getattr(cell, "col_idx", None)
            if column_index is None:
                column = cell.column
                column_index = column if isinstance(column, int) else column_index_from_string(column)

            target_cell = target.cell(row=cell.row, column=column_index)
            target_cell.value = cell.value
            target_cell.data_type = cell.data_type
            target_cell.number_format = cell.number_format
            target_cell.font = copy(cell.font)
            target_cell.fill = copy(cell.fill)
            target_cell.border = copy(cell.border)
            target_cell.alignment = copy(cell.alignment)
            target_cell.protection = copy(cell.protection)
            if cell.comment is not None:
                target_cell.comment = copy(cell.comment)
            if cell.hyperlink is not None:
                target_cell.hyperlink = copy(cell.hyperlink)

    for merged_range in source.merged_cells.ranges:
        target.merge_cells(str(merged_range))

    for row_idx, row_dimension in source.row_dimensions.items():
        target_dimension = target.row_dimensions[row_idx]
        target_dimension.height = row_dimension.height
        target_dimension.hidden = row_dimension.hidden
        target_dimension.outlineLevel = row_dimension.outlineLevel

    for column_key, column_dimension in source.column_dimensions.items():
        target_dimension = target.column_dimensions[column_key]
        target_dimension.width = column_dimension.width
        target_dimension.hidden = column_dimension.hidden
        target_dimension.outlineLevel = column_dimension.outlineLevel

    if source.conditional_formatting:
        for conditional_range, rules in source.conditional_formatting._cf_rules.items():
            for rule in rules:
                target.conditional_formatting.add(str(conditional_range.sqref), deepcopy(rule))

    parent = target.parent
    existing_table_names: set[str] = set()
    if parent is not None:
        for worksheet in parent.worksheets:
            existing_table_names.update(
                table.displayName for table in worksheet.tables.values() if table.displayName
            )

    for table in source.tables.values():
        new_table = deepcopy(table)
        base_name = table.displayName or table.name or "Table"
        candidate = base_name
        suffix = 1
        while candidate in existing_table_names:
            candidate = f"{base_name}_{suffix}"
            suffix += 1
        existing_table_names.add(candidate)
        new_table.displayName = candidate
        new_table.name = candidate
        target.add_table(new_table)

    target.HeaderFooter = deepcopy(source.HeaderFooter)

    target.sheet_properties.tabColor = (
        copy(source.sheet_properties.tabColor) if source.sheet_properties.tabColor else None
    )

    if source.auto_filter and (
        source.auto_filter.ref or source.auto_filter.filterColumn or source.auto_filter.sortState
    ):
        target.auto_filter = deepcopy(source.auto_filter)


def _parse_args(args: Sequence[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Merge multiple Excel workbooks into a single file where each file "
            "becomes its own worksheet."
        )
    )
    parser.add_argument(
        "source_directory",
        nargs="?",
        type=Path,
        help="Folder containing the Excel files to merge",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help=(
            "Path to the consolidated Excel workbook. Default: 'combined.xlsx' in "
            "the current directory (or in the script directory when no source "
            "folder is provided)."
        ),
    )
    parser.add_argument(
        "-p",
        "--pattern",
        default="*.xlsx",
        help="Glob pattern used to filter Excel files (default: '*.xlsx')",
    )
    parser.add_argument(
        "-r",
        "--recursive",
        action="store_true",
        help="Search for Excel files recursively",
    )
    parser.add_argument(
        "--values-only",
        action="store_true",
        help="Copy the last calculated values instead of formulas",
    )
    return parser.parse_args(args)


def _resolve_execution_paths(
    args: argparse.Namespace, script_path: Path
) -> tuple[Path, Path, str | None]:
    script_directory = script_path.parent
    default_output_name = Path("combined.xlsx")

    if args.source_directory is not None:
        source_directory = args.source_directory
        output_path = args.output or default_output_name
        message = None
    else:
        candidate: Path | None = DEFAULT_SOURCE_DIRECTORY
        if candidate is not None and not candidate.is_absolute():
            candidate = (script_directory / candidate).resolve()

        if candidate is not None and candidate.exists():
            source_directory = candidate
            message = (
                "No source directory provided. Using the default folder:\n"
                f"  {source_directory}\n"
            )
        else:
            source_directory = script_directory
            message = (
                "No source directory provided. Using the folder that contains this "
                f"script:\n  {source_directory}\n"
            )

        output_path = args.output or default_output_name
        if not output_path.is_absolute():
            output_path = script_directory / output_path

    return source_directory, output_path, message


def main(argv: Sequence[str] | None = None) -> int:
    if argv is None:
        argv = sys.argv[1:]
    args = _parse_args(argv)
    script_path = Path(__file__).resolve()
    source_directory, output_path, info_message = _resolve_execution_paths(args, script_path)
    if info_message:
        sys.stdout.write(info_message)
    try:
        merged = merge_excel_files(
            source_directory=source_directory,
            output_path=output_path,
            pattern=args.pattern,
            recursive=args.recursive,
            values_only=args.values_only,
        )
    except Exception as exc:  # pragma: no cover - CLI error handling
        message = f"Error: {exc}\n"
        sys.stderr.write(message)
        return 1

    file_count = len(merged)
    plural = "s" if file_count != 1 else ""
    sys.stdout.write(
        f"Merged {file_count} Excel file{plural} into '{output_path}'.\n"
    )
    return 0


if __name__ == "__main__":  # pragma: no cover - CLI entry point
    sys.exit(main())