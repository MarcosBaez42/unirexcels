from datetime import datetime
from pathlib import Path
import sys


ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

import merge_excel_files as merge_module
from merge_excel_files import merge_excel_files


def _create_workbook(path: Path, rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def test_merge_excel_files_creates_sheet_per_file(tmp_path: Path) -> None:
    first = tmp_path / "alpha.xlsx"
    second = tmp_path / "bravo.xlsx"
    _create_workbook(first, [["header", "value"], ["row", 1]])
    _create_workbook(second, [["only", "row"]])

    output = tmp_path / "combined.xlsx"
    merged = merge_excel_files(tmp_path, output)

    assert [entry.sheet_name for entry in merged] == ["alpha", "bravo"]

    workbook = load_workbook(output)
    try:
        assert workbook.sheetnames == ["alpha", "bravo"]
        alpha_rows = [tuple(row) for row in workbook["alpha"].iter_rows(values_only=True)]
        bravo_rows = [tuple(row) for row in workbook["bravo"].iter_rows(values_only=True)]
    finally:
        workbook.close()

    assert alpha_rows == [("header", "value"), ("row", 1)]
    assert bravo_rows == [("only", "row")]


def test_merge_excel_files_handles_duplicate_names(tmp_path: Path) -> None:
    long_name = "a" * 40
    first = tmp_path / f"{long_name}1.xlsx"
    second = tmp_path / f"{long_name}2.xlsx"
    _create_workbook(first, [[1]])
    _create_workbook(second, [[2]])

    output = tmp_path / "out.xlsx"
    merged = merge_excel_files(tmp_path, output)

    sheet_names = [entry.sheet_name for entry in merged]
    assert sheet_names[0] == "a" * 31
    assert sheet_names[1].startswith("a" * 27)
    assert sheet_names[1].endswith("_2")

    workbook = load_workbook(output)
    try:
        first_value = workbook[sheet_names[0]].cell(1, 1).value
        second_value = workbook[sheet_names[1]].cell(1, 1).value
    finally:
        workbook.close()

    assert first_value == 1
    assert second_value == 2


def test_merge_excel_files_searches_nested_directories_when_needed(tmp_path: Path) -> None:
    base_dir = tmp_path / "mis_excels" / "Reportes"
    nested_dir = base_dir / "Region" / "Daily"
    nested_dir.mkdir(parents=True)
    nested_file = nested_dir / "ventas.xlsx"
    _create_workbook(nested_file, [["Ventas"], [100]])

    output = base_dir / "combined.xlsx"
    merged = merge_excel_files(base_dir, output)

    assert [entry.source for entry in merged] == [nested_file]
    assert [entry.sheet_name for entry in merged] == ["ventas"]

    workbook = load_workbook(output)
    try:
        assert workbook.sheetnames == ["ventas"]
    finally:
        workbook.close()


def test_merge_excel_files_fails_when_no_files(tmp_path: Path) -> None:
    output = tmp_path / "merged.xlsx"
    with pytest.raises(ValueError):
        merge_excel_files(tmp_path, output)


def test_merge_excel_files_preserves_styles_and_features(tmp_path: Path) -> None:
    source = tmp_path / "styled.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Styled"

    sheet["A1"] = "Name"
    sheet["B1"] = "Value"
    sheet["C1"] = "Formula"
    sheet["A2"] = "Alpha"
    sheet["A2"].font = Font(bold=True)
    sheet["A2"].hyperlink = "http://example.com"
    sheet["B2"] = 123.456
    sheet["B2"].number_format = "#,##0.00"
    sheet["C2"] = "=B2*2"
    sheet["C2"].alignment = Alignment(horizontal="center")
    sheet["C2"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    sheet["C2"].border = border
    sheet["C2"].protection = Protection(locked=False)
    sheet["C2"].comment = Comment("Calculated value", "Tester")

    sheet.row_dimensions[1].height = 25
    sheet.row_dimensions[2].hidden = True
    sheet.row_dimensions[2].outlineLevel = 1
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].hidden = True
    sheet.column_dimensions["C"].outlineLevel = 2

    sheet.merge_cells("A3:C3")
    sheet["A3"] = "Merged"

    rule_fill = PatternFill(fill_type="solid", fgColor="00FF00")
    rule = CellIsRule(operator="equal", formula=["123.456"], fill=rule_fill, stopIfTrue=True)
    sheet.conditional_formatting.add("B2", rule)

    table = Table(displayName="StyledTable", ref="A1:C2")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    sheet.add_table(table)

    sheet.sheet_properties.tabColor = "1072BA"
    sheet.oddHeader.center.text = "Header Title"
    sheet.auto_filter.ref = "A1:C2"

    workbook.save(source)

    output = tmp_path / "combined.xlsx"
    merge_excel_files(tmp_path, output)

    merged = load_workbook(output)
    try:
        merged_sheet = merged["styled"]
        assert merged_sheet["A2"].font.bold is True
        assert merged_sheet["B2"].number_format == "#,##0.00"
        assert merged_sheet["C2"].value == "=B2*2"
        assert merged_sheet["C2"].data_type == "f"
        assert merged_sheet["C2"].alignment.horizontal == "center"
        assert merged_sheet["C2"].fill.fgColor.rgb == "00FFFF00"
        assert merged_sheet["C2"].border.left.style == "thin"
        assert merged_sheet["C2"].protection.locked is False
        assert merged_sheet["C2"].comment.text == "Calculated value"
        assert merged_sheet["A2"].hyperlink.target == "http://example.com"

        assert merged_sheet.row_dimensions[1].height == pytest.approx(25)
        assert merged_sheet.row_dimensions[2].hidden is True
        assert merged_sheet.row_dimensions[2].outlineLevel == 1
        assert merged_sheet.column_dimensions["A"].width == pytest.approx(20)
        assert merged_sheet.column_dimensions["B"].hidden is True
        assert merged_sheet.column_dimensions["C"].outlineLevel == 2

        merged_ranges = {str(rng) for rng in merged_sheet.merged_cells.ranges}
        assert "A3:C3" in merged_ranges

        cf_ranges = {str(cf.sqref) for cf in merged_sheet.conditional_formatting._cf_rules}
        assert "B2" in cf_ranges
        cf_rules = list(merged_sheet.conditional_formatting._cf_rules.values())[0]
        assert cf_rules[0].type == "cellIs"

        tables = list(merged_sheet.tables.values())
        assert len(tables) == 1
        assert tables[0].ref == "A1:C2"
        assert tables[0].tableStyleInfo.name == "TableStyleMedium2"
        assert tables[0].displayName.startswith("StyledTable")

        assert merged_sheet.sheet_properties.tabColor.rgb == "001072BA"
        assert merged_sheet.oddHeader.center.text == "Header Title"
        assert merged_sheet.auto_filter.ref == "A1:C2"
    finally:
        merged.close()


def test_main_defaults_to_script_directory(monkeypatch, tmp_path: Path) -> None:
    calls: dict[str, object] = {}

    def fake_merge(
        *,
        source_directory: Path,
        output_path: Path,
        pattern: str,
        recursive: bool,
        values_only: bool,
    ) -> list[merge_module.MergedSheet]:
        calls["source_directory"] = source_directory
        calls["output_path"] = output_path
        return []

    monkeypatch.setattr(merge_module, "merge_excel_files", fake_merge)
    script_file = tmp_path / "merge_excel_files.py"
    script_file.write_text("print('placeholder')\n")
    monkeypatch.setattr(merge_module, "__file__", str(script_file))

    exit_code = merge_module.main([])

    assert exit_code == 0
    assert calls["source_directory"] == tmp_path
    assert calls["output_path"] == tmp_path / "combined.xlsx"


def test_main_uses_cli_arguments(monkeypatch, tmp_path: Path) -> None:
    recorded: dict[str, object] = {}

    def fake_merge(
        *,
        source_directory: Path,
        output_path: Path,
        pattern: str,
        recursive: bool,
        values_only: bool,
    ) -> list[merge_module.MergedSheet]:
        recorded["source_directory"] = source_directory
        recorded["output_path"] = output_path
        recorded["pattern"] = pattern
        recorded["recursive"] = recursive
        recorded["values_only"] = values_only
        return [merge_module.MergedSheet(source=Path("alpha.xlsx"), sheet_name="alpha")]

    monkeypatch.setattr(merge_module, "merge_excel_files", fake_merge)

    args = [str(tmp_path), "--output", "out.xlsx", "--pattern", "*.xlsm", "--recursive", "--values-only"]
    exit_code = merge_module.main(args)

    assert exit_code == 0
    assert recorded["source_directory"] == tmp_path
    assert recorded["output_path"] == Path("out.xlsx")
    assert recorded["pattern"] == "*.xlsm"
    assert recorded["recursive"] is True
    assert recorded["values_only"] is True


def test_merge_excel_files_supports_xls(tmp_path: Path) -> None:
    fixture = Path(__file__).resolve().parent / "fixtures" / "sample.xls"
    destination = tmp_path / "sample.xls"
    destination.write_bytes(fixture.read_bytes())

    output = tmp_path / "combined.xlsx"
    merged = merge_excel_files(tmp_path, output)

    assert [entry.source.name for entry in merged] == ["sample.xls"]
    assert [entry.sheet_name for entry in merged] == ["sample"]

    workbook = load_workbook(output)
    try:
        sheet = workbook["sample"]
        assert sheet["A1"].value == "Name"
        assert sheet["B2"].value == pytest.approx(123.45)
        assert sheet["C2"].value == datetime(2023, 1, 15)
        merged_ranges = {str(rng) for rng in sheet.merged_cells.ranges}
        assert "A3:C3" in merged_ranges
    finally:
        workbook.close()